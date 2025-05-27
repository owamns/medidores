import pandas as pd
import re
import os
from datetime import datetime
from process_images import MeterProcessor


def extract_suministro_sucursal(filename):
    name_without_ext = os.path.splitext(filename)[0]

    if ' - ' in name_without_ext:
        parts = name_without_ext.split(' - ', 1)
        sucursal = parts[0].strip()
        suministro_part = parts[1].strip()
    else:
        sucursal = ""
        suministro_part = name_without_ext

    numeros = re.findall(r'\d+', suministro_part)
    suministro = ''.join(numeros) if numeros else ""
    suministro = suministro.replace("'", "")
    print(suministro)

    return int(suministro), sucursal


def generate_excel_from_images():
    crop_model_path = './model_recorte.pt'
    digital_model_path = 'models/best-digital.pt'
    analog_model_path = './models/best-electronico.pt'
    input_folder = './images'

    print("Iniciando procesamiento de imágenes...")

    processor = MeterProcessor(
        crop_model_path,
        digital_model_path,
        analog_model_path,
        input_folder,
        scale_factor=4
    )

    results = processor.process_images(show=False)

    if not results:
        print("No se encontraron resultados para procesar.")
        return

    print(f"Se procesaron {len(results)} imágenes.")

    excel_data = []

    for result in results:
        filename = result['image']
        lectura = float(result['detected_number'])

        suministro, sucursal = extract_suministro_sucursal(filename)

        excel_data.append({
            'suministro': suministro,
            'sucursal': sucursal,
            'lectura': lectura
        })

        print(f"Procesado: {filename} -> Suministro: {suministro}, Sucursal: {sucursal}, Lectura: {lectura}")

    df = pd.DataFrame(excel_data)

    df = df.sort_values(['sucursal', 'suministro']).reset_index(drop=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_filename = f"lecturas_medidores_{timestamp}.xlsx"

    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Lecturas', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Lecturas']

        worksheet.column_dimensions['A'].width = 15  # suministro
        worksheet.column_dimensions['B'].width = 25  # sucursal
        worksheet.column_dimensions['C'].width = 15  # lectura

        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 4):  # Columnas A, B, C
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        summary_data = {
            'Métrica': [
                'Total de lecturas',
                'Total de sucursales',
                'Total de suministros únicos',
                'Fecha de procesamiento'
            ],
            'Valor': [
                len(df),
                df['sucursal'].nunique(),
                df['suministro'].nunique(),
                datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Resumen', index=False)

        summary_worksheet = writer.sheets['Resumen']
        summary_worksheet.column_dimensions['A'].width = 25
        summary_worksheet.column_dimensions['B'].width = 20

        for col in range(1, 3):  # Columnas A, B
            cell = summary_worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    print(f"\n¡Excel generado exitosamente!")
    print(f"Archivo: {excel_filename}")
    print(f"Ubicación: {os.path.abspath(excel_filename)}")
    print(f"Total de registros: {len(df)}")

    return excel_filename


if __name__ == "__main__":
    try:
        generate_excel_from_images()
    except Exception as e:
        print(f"Error al generar el Excel: {str(e)}")
        print("Verifica que:")
        print("1. Los modelos (.pt) estén en las rutas correctas")
        print("2. La carpeta 'images' exista y contenga imágenes")
        print("3. Tengas instalado: pip install pandas openpyxl")